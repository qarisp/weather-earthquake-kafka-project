import requests
import pandas as pd
import sys
from datetime import datetime
from openpyxl import load_workbook

def extract(API_KEY, PROVINCES):

    weather_list = []
    BASE_URL = 'https://api.openweathermap.org/data/2.5/weather'

    for province in PROVINCES:
        params = {
            'lat': province['lat'],
            'lon': province['lon'],
            'appid': API_KEY,
            'units': 'metric'
        }

        response = requests.get(BASE_URL, params=params)

        if response.status_code == 200:
            data = response.json()
            weather_list.append({
                'Province': province['name'],
                'Capital City': province['capital_city'],
                'Temperature (C)': round(data['main']['temp'], 2),
                'Weather': data['weather'][0]['description'],
                'Humidity (%)': data['main']['humidity'],
                'Pressure (Pa)': data['main']['pressure'],
                'Wind Speed (m/s)': data['wind']['speed'],
                'lat': province['lat'],
                'lon': province['lon']
            })
        else:
            print(f"Error for {province['name']}: {response.status_code}")
            sys.exit(1)

    raw_data = pd.DataFrame(weather_list)

    return raw_data

def transform(raw_data):

    clean_data = raw_data.copy(deep=True)

    clean_data['Datetime'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    clean_data = clean_data[[
        'Datetime',
        'Province',
        'Capital City',
        'Temperature (C)',
        'Weather',
        'Humidity (%)',
        'Pressure (Pa)',
        'Wind Speed (m/s)',
        'lat', 'lon'
    ]]

    return clean_data

def load(clean_data):
    file_path = 'weather_data.xlsx'
    
    try:
        # Load existing workbook and find the last row
        book = load_workbook(file_path)
        sheet = book.active
        last_row = sheet.max_row

        # Create writer with append mode
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            clean_data.to_excel(writer, index=False, header=False, startrow=last_row)
    
    except FileNotFoundError:
        # Create file if it doesn't exist
        clean_data.to_excel(file_path, index=False)